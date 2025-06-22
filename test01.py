import time
from DrissionPage import ChromiumPage
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def load_comments(page, timeout=20):
    """加载评论，使用超时机制"""
    print('开始下滑加载主评论...')
    previous_count = 0
    last_growth_time = time.time()

    while True:
        for _ in range(3):
            page.scroll.down(800)
            time.sleep(1)

        threads = page.eles('xpath://ytd-comment-thread-renderer')
        current_count = len(threads)
        print(f'当前主评论数: {current_count}')

        if current_count > previous_count:
            last_growth_time = time.time()
            previous_count = current_count
        else:
            elapsed = time.time() - last_growth_time
            print(f'未检测到新评论，已等待 {int(elapsed)} 秒...')
            if elapsed >= timeout:
                print(f'{timeout}秒内没有新评论加载，停止下滑')
                break
            time.sleep(2)

def scroll_until_find(page, xpath, max_attempts=5):
    """下滑直到找到指定元素"""
    attempts = 0
    while attempts < max_attempts:
        elements = page.eles(f'xpath:{xpath}')
        if elements:
            return elements
        page.scroll.down(800)
        time.sleep(1)
        attempts += 1
    return []

def expand_replies(page):
    """展开所有评论的回复"""
    print('展开所有回复与更多按钮...')

    # 展开回复按钮
    while True:
        reply_btns = scroll_until_find(page, '//*[@id="more-replies"]/yt-button-shape/button')
        if not reply_btns:
            break

        clicked = False
        for btn in reply_btns:
            try:
                if btn.states.is_displayed:
                    btn.click()
                    time.sleep(1)
                    clicked = True
            except:
                continue
        if not clicked:
            break

    # 展开更多回复
    while True:
        more_btns = scroll_until_find(page, '//*[@id="button"]/ytd-button-renderer//button')
        if not more_btns:
            break

        clicked = False
        for btn in more_btns:
            try:
                if btn.states.is_displayed:
                    btn.click()
                    time.sleep(1)
                    clicked = True
            except:
                continue
        if not clicked:
            break

    # 展开长评论
    while True:
        expand_btns = scroll_until_find(page, '//*[@id="more"]/span')
        if not expand_btns:
            break

        clicked = False
        for btn in expand_btns:
            try:
                if btn.states.is_displayed:
                    btn.click()
                    time.sleep(0.5)
                    clicked = True
            except:
                continue
        if not clicked:
            break

    print('完成展开所有可见的回复')

def extract_comments(page):
    """提取评论数据"""
    print('提取评论数据...')
    all_data = []
    threads = page.eles('xpath://ytd-comment-thread-renderer')
    total_replies = 0

    for idx, thread in enumerate(threads, 1):
        try:
            # 提取主评论
            author = thread.ele('xpath:.//*[@id="author-text"]/span').text.strip()
            content = thread.ele('xpath:.//*[@id="content-text"]/span').text.strip()
            like = thread.ele('xpath:.//*[@id="vote-count-middle"]').text.strip() or '0'

            # 提取回复
            replies = thread.eles('xpath:.//ytd-comment-replies-renderer//*[@id="content-text"]/span')
            reply_authors = thread.eles('xpath:.//ytd-comment-replies-renderer//*[@id="author-text"]/span')
            reply_likes = thread.eles('xpath:.//ytd-comment-replies-renderer//*[@id="vote-count-middle"]')

            reply_list = []
            for ra, rc, rl in zip(reply_authors, replies, reply_likes):
                reply_list.append([ra.text.strip(), rc.text.strip(), rl.text.strip() or '0'])

            total_replies += len(reply_list)
            all_data.append([author, content, like, 0, reply_list])
            print(f'已提取主评论{idx}及其{len(reply_list)}条回复')

        except Exception as e:
            print(f'提取主评论{idx}时出错: {str(e)}')
            continue

    print(f'总计提取：{len(all_data)}个主评论，{total_replies}个回复')
    return all_data

def save_to_excel(data, filename="YouTube评论.xlsx"):
    """保存到Excel"""
    print('写入Excel...')
    wb = Workbook()
    ws = wb.active
    ws.title = "YouTube评论结构"

    # 写入表头
    ws.append(["序号", "账号", "内容", "赞", "踩", "序号", "账号", "内容", "赞"])

    # 写入数据
    row_index = 2
    for i, item in enumerate(data, 1):
        author, content, like, dislike, replies = item
        start_row = row_index

        if replies:
            for j, reply in enumerate(replies, 1):
                ws.append([
                    i if j == 1 else "",
                    author if j == 1 else "",
                    content if j == 1 else "",
                    like if j == 1 else "",
                    dislike if j == 1 else "",
                    j, reply[0], reply[1], reply[2]
                ])
                row_index += 1
            # 合并主评论单元格
            for col in range(1, 6):
                ws.merge_cells(start_row=start_row, end_row=row_index-1, start_column=col, end_column=col)
                ws.cell(row=start_row, column=col).alignment = Alignment(horizontal='center', vertical='center')
        else:
            ws.append([i, author, content, like, dislike, '', '', '', ''])
            row_index += 1

    # 自动调整列宽
    for col in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = length + 2

    wb.save(filename)
    print(f'✅ 完成！文件已保存为：{filename}')

def main():
    url = 'https://www.youtube.com/watch?v=6ttRK5OMxf4'
    print('启动浏览器并打开页面...')
    page = ChromiumPage()
    page.get(url)

    page.scroll.down(400)
    time.sleep(2)
    
    try:
        load_comments(page)  # 加载评论
        expand_replies(page)  # 展开回复
        all_data = extract_comments(page)  # 提取数据
        save_to_excel(all_data)  # 保存到Excel
    finally:
        page.close()

if __name__ == '__main__':
    main()
